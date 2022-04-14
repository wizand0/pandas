import collections
from collections import defaultdict, Counter
import pandas
from openpyxl import load_workbook

NUMBER_OF_POPULAR_BROWERS = 7
NUMBER_OF_POPULAR_GOODS = 7


def make_report(log_file_name, report_template_file_name, report_output_file_name):
    # Чтение и анализ данных из Excel
    excel_data = pandas.read_excel(log_file_name, sheet_name='log', engine='openpyxl')
    excel_data_dict = excel_data.to_dict(orient='records')

    visits_dict = defaultdict(int)
    goods_dict = []
    visits_dict_month = {}
    popular_goods_month = {}

    m_goods = []
    f_goods = []
    for element in excel_data_dict:
        visits_dict[element['Браузер']] += 1

        if element['Купленные товары']:
            temp_goods = element['Купленные товары'].split(',')
            for el in temp_goods:
                goods_dict.append(el)

        if element['Пол'] == 'м':
            temp_goods = element['Купленные товары'].split(',')
            for el in temp_goods:
                m_goods.append(el)
        if element['Пол'] == 'ж':
            temp_goods = element['Купленные товары'].split(',')
            for el in temp_goods:
                f_goods.append(el)

    # Ищем самые популярные браузеры и товары и самые непопулярные
    most_popular_browsers = Counter(visits_dict).most_common(NUMBER_OF_POPULAR_BROWERS)
    most_popular_goods = collections.Counter(goods_dict).most_common(NUMBER_OF_POPULAR_GOODS)
    most_popular_m_goods = collections.Counter(m_goods).most_common(1)
    temp_less_popular_m_goods = collections.Counter(m_goods).most_common()
    less_popular_m_goods = temp_less_popular_m_goods[:-(len(temp_less_popular_m_goods) + 1):-1][0]
    most_popular_f_goods = collections.Counter(f_goods).most_common(1)
    temp_less_popular_f_goods = collections.Counter(f_goods).most_common()
    less_popular_f_goods = temp_less_popular_f_goods[:-(len(temp_less_popular_f_goods) + 1):-1][0]

    for element in excel_data_dict:
        datestamp = element['Дата посещения']
        date1 = datestamp.to_pydatetime()
        date2 = date1.date()
        number_of_month = int(date2.strftime("%m"))

        for i in range(NUMBER_OF_POPULAR_BROWERS):
            if element['Браузер'] == str(most_popular_browsers[i][0]):
                if str(most_popular_browsers[i][0]) in visits_dict_month:
                    for m in range(0, 11):
                        if number_of_month == m:
                            if number_of_month in visits_dict_month[str(most_popular_browsers[i][0])]:
                                visits_dict_month[str(most_popular_browsers[i][0])][number_of_month] += 1
                            else:
                                visits_dict_month[str(most_popular_browsers[i][0])][number_of_month] = 1
                else:
                    visits_dict_month[str(most_popular_browsers[i][0])] = {number_of_month: 1}

    for element in excel_data_dict:
        datestamp = element['Дата посещения']
        date1 = datestamp.to_pydatetime()
        date2 = date1.date()
        number_of_month = int(date2.strftime("%m"))

        temp_goods_month = element['Купленные товары'].split(',')

        for elem in temp_goods_month:
            for i in range(0, NUMBER_OF_POPULAR_GOODS):
                for item in most_popular_goods[i]:
                    if elem == item:
                        if elem in popular_goods_month:
                            if number_of_month in popular_goods_month[elem]:
                                popular_goods_month[elem][number_of_month] += 1
                            else:
                                popular_goods_month[elem][number_of_month] = 1
                        else:
                            popular_goods_month[elem] = {}
                            popular_goods_month[elem][number_of_month] = 1

    # Открываем файл шаблона отчета report_template.xlsx
    wb = load_workbook(filename=report_template_file_name)
    ws = wb.active

    # Заполняем таблицу по использованию браузеров
    # В этом цикле заполняем популярные браузеры. Количество нормируется константой
    for i in range(1, NUMBER_OF_POPULAR_BROWERS + 1):
        int_row = ord('A')
        row = chr(int_row) + str(5 + i - 1)
        ws[row] = str(most_popular_browsers[i - 1][0])
        # Во вложенном цикле заполняем посещяемость по месяцам
        for j in range(1, 12):
            int_col = ord('A')
            cell = chr(int_col + j) + str(5 + i - 1)
            try:
                ws[cell] = str(visits_dict_month[str(most_popular_browsers[i - 1][0])][j])
            except:
                pass

    # Заполняем таблицу по приобретенным товарам
    # В этом цикле заполняем популярные товаров. Количество нормируется константой
    for i in range(1, NUMBER_OF_POPULAR_GOODS + 1):
        int_row = ord('A')
        row = chr(int_row) + str(19 + i - 1)
        ws[row] = str(most_popular_goods[i - 1][0])
        # pprint(most_popular_goods[i - 1][0])
        # Во вложенном цикле заполняем покупки популярных товаров по месяцам
        for j in range(1, 12):
            int_col = ord('A')
            cell = chr(int_col + j) + str(19 + i - 1)
            try:
                # Поиск ключа (товара, для которого нужно проставить количество продаж
                temp_key = str(most_popular_goods[i - 1][0])
                ws[cell] = str(popular_goods_month[temp_key][j])
            except:
                pass

    # Заполняем самые популярные и непопулярные товары у мужчин и женщин
    ws['B31'] = str(most_popular_m_goods[0][0])
    ws['B32'] = str(most_popular_f_goods[0][0])
    ws['B33'] = str(less_popular_m_goods[0])
    ws['B34'] = str(less_popular_f_goods[0])

    # Сохраняем файл-отчет
    wb.save(report_output_file_name)

